from tkinter import *
import openpyxl
from openpyxl.chart import Reference
from openpyxl.chart import PieChart
from win32com import client

root = Tk()
# set the width and height of the window
root.geometry("400x400")
# set the fixed size of window
root.resizable(0, 0)
# set title of window
root.title("Test Cases parser")
Label(root, text="Test Cases Parser", font='arial 15 bold').pack()

# create label and entry for tester name
tester_label = Label(root, text='Tester Name:', font='arial 10 bold').pack()
tester_str = StringVar()
Entry(root, textvariable=tester_str).pack()

path = r'C:\Users\Picasso\Desktop\Curs Software Tester\Proiect\Test_Case_Format_Farcas_Daniel.xlsx'
pathPDF = r"C:\Users\Picasso\Desktop\Curs Software Tester\Proiect\Test_Case_Format_Farcas_Daniel.pdf"

values = [0, 0]


def generateReport():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb['TestCaseFormat']
    global tester_str
    tester_str = sheet.cell(1, 5).value

    try:
        reportSheet = wb['Report']
    except:
        wb.create_sheet('Report')
        reportSheet = wb['Report']

    reportSheet['A1'] = 'Tester ID:'
    reportSheet['B1'] = tester_str

    reportSheet['A2'] = 'Failed test cases'
    reportSheet['B2'] = values[1]

    reportSheet['A3'] = 'Passed test cases'
    reportSheet['B3'] = values[0]

    reportSheet['A4'] = 'Total'
    reportSheet['B4'] = values[0] + values[1]

    wb.save(path)
    createChart()

    # Open Microsoft Excel
    excel = client.Dispatch('Excel.Application')

    # Read Excel File
    sheets = excel.Workbooks.Open(path)
    work_sheets = sheets.Worksheets[2]

    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, pathPDF)

def compareValues():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb['TestCaseFormat']

    for row in range(1, int(sheet.max_row + 1)):

        if sheet.cell(row=row, column=7).value == 'Pass':
            values[0] = values[0] + 1
        elif sheet.cell(row=row, column=7).value == 'Fail':
            values[1] = values[1] + 1


def createChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb['Report']
    pie = PieChart()

    labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(sheet, min_col=2, min_row=2, max_row=3)
    pie.add_data(data,titles_from_data=False)
    pie.set_categories(labels)
    pie.title='Test Cases'

    pie.width = 14
    pie.height = 7
    s = pie.series[0]
    s.graphicalProperties.line.solidFill = "00000"
    sheet.add_chart(pie, 'B7')

    wb.save(path)

def buttonPressed():
    compareValues()
    generateReport()

Button(root, text="Generate Report", command=buttonPressed).pack(pady=5)

root.mainloop()
